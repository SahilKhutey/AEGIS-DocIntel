"""
AEGIS-AMDI-OS — XLSX Loader
=============================
Microsoft Excel spreadsheet loader using openpyxl.
"""
from __future__ import annotations

import io
import logging
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from src.core.document_object import DocumentFormat, DocumentObject
from src.ingestion.base import BaseLoader, FormatError, LoaderError, SizeLimitError

logger = logging.getLogger(__name__)


class XLSXLoader(BaseLoader):
    """XLSX spreadsheet loader."""

    FORMAT_NAME = "xlsx"
    SUPPORTED_EXTENSIONS = {".xlsx"}
    XLSX_MAGIC = b"PK\x03\x04"

    def __init__(self, max_size_mb: int = 100, **options):
        super().__init__(**options)
        self.max_size_mb = max_size_mb

    def validate(self, raw_bytes: bytes) -> bool:
        if not raw_bytes or len(raw_bytes) < 4:
            return False
        return raw_bytes[:4] == self.XLSX_MAGIC

    async def load(self, source, filename: str = "") -> DocumentObject:
        raw_bytes, name = self.read_source(source)
        if filename:
            name = filename
        if not name:
            name = "spreadsheet.xlsx"

        if not self.validate(raw_bytes):
            raise FormatError("Not a valid XLSX file")

        size_mb = len(raw_bytes) / (1024 * 1024)
        if size_mb > self.max_size_mb:
            raise SizeLimitError(f"XLSX too large: {size_mb:.1f}MB")

        metadata, text_parts, sheet_count = self._extract(raw_bytes)
        return DocumentObject(
            filename=name,
            format=DocumentFormat.XLSX,
            raw_bytes=raw_bytes,
            metadata=metadata,
            page_count=sheet_count,
            text_content="\n\n".join(text_parts),
        )

    def _extract(self, raw_bytes: bytes) -> tuple[dict[str, Any], list[str], int]:
        metadata: dict[str, Any] = {}
        text_parts: list[str] = []
        sheet_count = 0
        try:
            wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
            metadata["sheet_count"] = len(wb.sheetnames)
            metadata["sheet_names"] = wb.sheetnames
            metadata["creator"] = wb.properties.creator
            metadata["title"] = wb.properties.title
            metadata["subject"] = wb.properties.subject
            metadata["keywords"] = wb.properties.keywords
            metadata["created"] = str(wb.properties.created) if wb.properties.created else None
            metadata["modified"] = str(wb.properties.modified) if wb.properties.modified else None

            sheet_count = len(wb.sheetnames)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                row_count = 0
                col_count = 0
                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    col_count = max(col_count, sum(1 for c in row if c is not None))
                    row_data = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in row_data):
                        text_parts.append(" | ".join(row_data))
                metadata[f"sheet_{sheet_name}_rows"] = row_count
                metadata[f"sheet_{sheet_name}_cols"] = col_count
        except Exception as e:
            logger.warning(f"XLSX extraction failed: {e}")
        return metadata, text_parts, sheet_count
